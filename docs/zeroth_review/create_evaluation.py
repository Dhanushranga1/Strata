#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Create workbook
wb = openpyxl.Workbook()

# Create Rubric worksheet
ws_rubric = wb.active
ws_rubric.title = "Rubric"

# Headers
headers = ["CO_Range", "CriterionID", "CriterionName", "Category", "Description", "MaxScore", "ScoreGiven", "Comments"]
ws_rubric.append(headers)

# Style headers
for col in range(1, len(headers) + 1):
    cell = ws_rubric.cell(1, col)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Rubric data
rubric_data = [
    ["CO 1 to 5", "2.1.1", "1. Problem Identification & Formulation - IA", "IA", "Clarity, relevance, and feasibility of the problem statement; alignment with the project domain.", 2, "", ""],
    ["CO 1 to 5", "2.2.3", "2. Literature Survey (Background Study)/ Market and patent analysis - IA", "IA", "Depth and relevance of reviewed literature; ability to connect existing work to the problem domain and the identification of technical gaps", 2, "", ""],
    ["CO 1 to 5", "2.5.1", "4. Research Objectives/ Product goals Formulation - GA", "GA", "Clarity, measurability, and relevance of project objectives/goals derived from the problem and gap analysis.", 2, "", ""],
    ["CO 1 to 5", "7.3.2", "5. Alignment with Sustainable Development Goals (SDGs) - GA", "GA", "Extent to which the project aligns with relevant UN SDGs and demonstrates societal relevance.", 2, "", ""],
    ["CO 1 to 5", "10.5.2", "6. Presentation & Documentation Quality - IA", "IA", "Clarity, organization, and professionalism of the presentation and documentation.", 2, "", ""]
]

for row in rubric_data:
    ws_rubric.append(row)

# Add summary rows
ws_rubric.append([])  # Empty row
ws_rubric.append(["", "", "Total IA Score:", "=SUMIFS(G:G,D:D,\"IA\")", "", "", "", ""])
ws_rubric.append(["", "", "Total GA Score:", "=SUMIFS(G:G,D:D,\"GA\")", "", "", "", ""])
ws_rubric.append([])  # Empty row
ws_rubric.append(["", "", "Decision:", "=IF(SUMIFS(G:G,D:D,\"IA\")>=5,\"Accepted\",IF(AND(SUMIFS(G:G,D:D,\"GA\")>2.5,SUMIFS(G:G,D:D,\"GA\")<5),\"Accepted with Revision\",\"Rejected\"))", "", "", "", ""])

# Adjust column widths
ws_rubric.column_dimensions['A'].width = 12
ws_rubric.column_dimensions['B'].width = 12
ws_rubric.column_dimensions['C'].width = 60
ws_rubric.column_dimensions['D'].width = 10
ws_rubric.column_dimensions['E'].width = 80
ws_rubric.column_dimensions['F'].width = 10
ws_rubric.column_dimensions['G'].width = 12
ws_rubric.column_dimensions['H'].width = 40

# Create ScoringGuide worksheet
ws_guide = wb.create_sheet("ScoringGuide")
ws_guide.append(["Performance Level", "Numeric Score", "Description"])

# Style headers
for col in range(1, 4):
    cell = ws_guide.cell(1, col)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Scoring guide data
guide_data = [
    ["Excellent", 2, "Outstanding work; exceeds all expectations; comprehensive and well-executed"],
    ["Good", 1.5, "Very good work; meets all expectations with minor areas for improvement"],
    ["Satisfactory", 1, "Acceptable work; meets minimum requirements; some gaps or weaknesses"],
    ["Needs Improvement", 0.5, "Below expectations; significant gaps or weaknesses; requires substantial revision"],
    ["Unsatisfactory", 0, "Does not meet requirements; major deficiencies; unacceptable quality"]
]

for row in guide_data:
    ws_guide.append(row)

# Adjust column widths
ws_guide.column_dimensions['A'].width = 20
ws_guide.column_dimensions['B'].width = 15
ws_guide.column_dimensions['C'].width = 80

# Save workbook
wb.save("evaluation.xlsx")
print("evaluation.xlsx created successfully")
